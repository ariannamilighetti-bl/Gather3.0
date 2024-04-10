# -*- coding: utf-8 -*-
"""
Created on Tue Jan 30 14:10:39 2024

@author: amilighe
"""
from tkinter import *
import tkinter as tk
from tkinter import filedialog
from datetime import datetime

import openpyxl.cell
from lxml import etree
from lxml.builder import ElementMaker
from lxml.etree import Comment
from openpyxl import load_workbook

# These are the column number for the fields used
# If the template changes, change the column numbers here
# IAMS template:
# Identification
repository_clmn = 0
coll_area_clmn = 1
collection_clmn = 2
level_clmn = 3
reference_clmn = 4
ext_ref_clmn = 5
# Title Information
title_clmn = 6
# Date and Calendar
date_rng_clmn = 7
start_date_clmn = 8
end_date_clmn = 9
era_clmn = 10
calendar_clmn = 11
# Content and Use
extent_clmn = 12
scope_content_clmn = 13
phys_char_clmn = 14
# Conditions of Access and Use
access_cond_clmn = 15
# Context
imm_acq_column = 16
cust_hist_clmn = 17
admin_context_clmn = 18
arrangement_clmn = 19
# Allied Material
rel_mat_clmn = 20
find_aids_clmn = 21
or_info_clmn = 22
copise_info_clmn = 23
pub_notes_clmn = 24
exhib_clmn = 25
rel_arch_des_clmn = 26
# Languages and Scripts
mat_language_clmn = 27
mat_langcode_clmn = 28
mat_script_clmn = 29
mat_scriptcode_clmn = 30
descr_lang_clmn = 31
descr_langcode_clmn = 32
descr_script_clmn = 33
descr_scriptcode_clmn = 34
# Authority Relationships
rel_persons_clmn = 35
rel_fams_clmn = 36
rel_corp_bds_clmn = 37
rel_places_clmn = 38
rel_subject_clmn = 39
# Cartographic
dec_lat_clmn = 40
dec_long_clmn = 41
dec_coords_clmn = 42
scale_clmn = 43
scale_des_clmn = 44
projection_clmn = 45
orientation_clmn = 46
# Record Control
legal_sts_clmn = 47
lev_details_clmn = 48
Visibility_clmn = 49
log_type_clmn = 50
log_lab_clmn = 51
page_seq_range = 52
page_lab_range = 53
mat_type_clmn = 54
item_type_clmn = 55
# Identifiers
ark_id_clmn = 56
iams_id_clmn = 57


# Gather definitions
# Definitions used to create the nodes:
def get_header(ws):
    '''Returns the values of the header row'''
    header = []
    for cell in ws[2]:
        header.append(cell.value)
    return header


def start_record(rec_num):
    '''Keeps count of how the record number in the file'''
    return {"StartRecord": rec_num}


def tid(row, arg, shelfmark_modified, row_num):
    '''Add the tid labels to the xml nodes'''
    # the tid_num should grow within the xml across child records
    global tid_num
    if row_num == 0:
        tid_num = 1
    if row[arg].value:
        tid_full = shelfmark_modified+"_"+str(tid_num)
        tid_num = tid_num+1
        return {"tid": tid_full}
    else:
        return {}


def content(row, arg):
    '''Adds the content of non-p nodes'''
    if row[arg].value:
        return str(row[arg].value)
    else:
        return {}


def labels(row, arg, label):
    '''Creates the labels of the xml nodes'''
    content = str(row[arg].value)
    return {label: content}


def header_label(header_row, arg, label):
    '''Returns the text of the column header'''
    label_title = header_row[arg]
    return {label: label_title}


def date_normal(row, arg):
    '''Returns the date field in the format start/end'''
    if row[start_date_clmn].value:
        if row[end_date_clmn].value:
            date = str(row[start_date_clmn].value)+"/"+str(row[end_date_clmn].value)
        else:
            date = str(row[start_date_clmn].value)+"/"+str(row[start_date_clmn].value)
    else:
        full_date = str(row[arg].value)
        if "-" in full_date:
            index = full_date.rfind("-")
            start_date = full_date[index-4:index]
            end_date = full_date[-4:]
            date = start_date+"/"+end_date
        else:
            date = full_date[-4:]+"/"+full_date[-4:]
    return {"normal": date}


"""
I've added type hints to this fn.
Python is dynamically typed - languages like C/Java are statically typed
This means you don't have to declare a type for each variable and code is more readable
But errors can occur if a var has the wrong type
Type hints explain what type each arg to a fn should have
This makes the fn easier to understand and IDEs can pick up when the wrong type is being passed in
"""
def pcontent(
        text: str,  # These are called type hints, they help explain what type each arg should be
        row: tuple[openpyxl.cell.ReadOnlyCell],
        arg: int,
        E: ElementMaker,
        shelfmark_modified: str,
        row_num: int,
 ) -> list[str]:  # can also add a type hint for the return val
    '''Creates the p node of free text fields and bullet point logic'''
    global tid_num
    content = []
    lists = E.list()
    if text:
        paragraph_initial = text.replace("<p><list>","<list>").strip()
        paragraphs = paragraph_initial.split("</p>")
        # print("paragraph =", paragraphs)
        for chunk in paragraphs:
            list_content = []
            lines = chunk.split("</item>")
            for line in lines:
                line = line.strip()
                if line != "":
                    tid_label = tid(row, arg, shelfmark_modified, row_num)['tid']  # we need the tid val now not a dict
                    if line.find("<emph render='italic'>") != -1:
                        line = line.replace("<emph render='italic'>",'<emph render="italic">')
                    if line.find('<emph render="italic">') != -1:
                        section1 = line.replace("</emph><emph","</emph> <emph")
                        sections = section1.split('<emph')
                        emphatic_line = []
                        for section in sections:
                            if section.find('render="italic">') != -1:
                                top = section.split(' render="italic">')[0]
                                emph_all = section.split(' render="italic">')[1]
                                emph = emph_all.split('</emph>')[0]
                                bottom = section.split("</emph>")[1]
                                emph_tid = shelfmark_modified+"_"+str(tid_num)
                                tid_num += 1
                                # E.emph is the equivalent here of E.subtag in the ElementMaker docs
                                # I've replaced the strings with a list including E.emph elements for any emphatics required
                                emphatic_line += [top]
                                emphatic_line += [E.emph(emph, render="italic", tid=emph_tid)]
                                emphatic_line += [bottom]
                            else:
                                emphatic_line += [section]
                            line = emphatic_line

                    # if no emphatics in the line we still need to make it a list for the next logic
                    if type(line) != list:
                        line = [line]

                    # need to iterate over the elements of line and replace certain strings with ""
                    # not all elements of line are strings - some are E.emph elements
                    # replace_if_str lets us safely skip these
                    def replace_if_str(elem, *args):
                        """
                        Replace all args with "" if elem is a string
                        Otherwise return elem
                        """
                        if type(elem) == str:
                            for a in args:
                                elem = elem.replace(a, "")
                            return elem
                        else:
                            return elem

                    # original logic was just str.replace
                    # have changed line to a list of str/elements so need to iterate this now
                    # this is true for all the str.replace calls that followed
                    line = [replace_if_str(elem, "<list>", "</list>") for elem in line]

                    if line[0].startswith("<item>"):
                        # line = line.replace("<item>", "")
                        line = [replace_if_str(elem, "<item>") for elem in line]
                        list_content.append(line)
                        line_content = E.item(*line, tid=tid_label)  # crucially need a `*` to unpack line
                        lists.append(line_content)
                        content.append(lists)
                    elif list_content == []:
                        line = [replace_if_str(elem, "<p>") for elem in line]
                        top_p = E.p(*line, tid=tid_label)  # crucially need a `*` to unpack line
                        content.append(top_p)
                    else:
                        line = [replace_if_str(elem, "<p>") for elem in line]
                        bttm_p = E.p(*line, tid=tid_label)  # crucially need a `*` to unpack line
                        content.append(bttm_p)
    else:
        p = E.p()
        content.append(p)
    return content


def title_content(row: tuple[openpyxl.cell.ReadOnlyCell],
        arg: int,
        E: ElementMaker,
        shelfmark_modified: str,
        row_num: int,
 ) -> list[str]:
    '''Creates the p node of free text fields and bullet point logic'''
    global tid_num
    if row[arg].value:
        tid_label = tid(row, arg, shelfmark_modified, row_num)['tid']
        line = row[arg].value
        if line.find("<emph render='italic'>") != -1:
            line = line.replace("<emph render='italic'>",'<emph render="italic">')
        if line.find('<emph render="italic">') != -1:
            section1 = line.replace("</emph><emph","</emph> <emph")
            sections = section1.split('<emph')
            emphatic_line = []
            for section in sections:
                if section.find('render="italic">') != -1:
                    top = section.split(' render="italic">')[0]
                    emph_all = section.split(' render="italic">')[1]
                    emph = emph_all.split('</emph>')[0]
                    bottom = section.split("</emph>")[1]
                    emph_tid = shelfmark_modified+"_"+str(tid_num)
                    tid_num += 1
                    # E.emph is the equivalent here of E.subtag in the ElementMaker docs
                    # I've replaced the strings with a list including E.emph elements for any emphatics required
                    emphatic_line += [top]
                    emphatic_line += [E.emph(emph, render="italic", tid=emph_tid)]
                    emphatic_line += [bottom]
                else:
                    emphatic_line += [section]
                line = emphatic_line
         # if no emphatics in the line we still need to make it a list for the next logic
                if type(line) != list:
                    line = [line]

                title_full = E.title(*line, tid=tid_label)  # crucially need a `*` to unpack line
                # content.append(title_full)
            return title_full
        else:
            title_full = E.title(line, tid=tid_label)
            return title_full
    else:
        title_full = E.title()
        return title_full


def current_wordcount(row):
    '''Calculates the number of workds in the Excel row'''
    wc = 0
    for i in range(0, len(row), 1):
        if row[i].value:
            par = str(row[i].value).replace("><", " ")
            wc += len(par.split())
    return wc


# Authority Files processing definitions
def auth_validation(auth_ws):
    header = []
    for cell in auth_ws[1]:
        header.append(cell.value)
    if len(header)>= 5 and str(header[0])=="Name":
        validated = True
    else:
        validated = False
    return validated

# Generate athority lookup
def gen_auth_lookup(auth_ws):
    '''Function looks at the Authority files sheet and creates a dictionary 
        of authority files in the format:
           auth_file_attr[0] = name
           auth_file_attr[1] = Ark id
           auth_file_attr[2] = role
           auth_file_attr[3] = Altrender
           auth_file_attr[4] = IAMS ID
        returns a lookup to the row number in the file.'''

    auth_lookup = {}
    reference = 1
    for row_num, row in enumerate(auth_ws.iter_rows(min_row=2)):
        possible_types = ["(Event)","(Term)","(Authorised)","(Parallel)","(Other)","(Building)","(Title of Work)"]
        reference += 1
        auth_file_attr = []
        single_auth_name = []
        all_names = row[0].value.replace("|", " xx ").replace(" xxx ", " xx ")
        if all_names.find(" xx ") != -1:
            names_and_types = all_names.split(" xx ")
        else:
            names_and_types = [all_names]
        for name in names_and_types:
            for find_type in possible_types:
                if find_type in name:
                    auth_name = name.split(find_type)[0]
                    auth_type = find_type.replace("(","").replace(")","")
                    break
                else:
                    auth_name = name
                    auth_type = "not_found"
            if auth_name:
                single_auth_name.append(auth_name)
            else:
                single_auth_name.append("not_found")
            if row[1].value:
                single_auth_name.append(row[1].value)
            else:
                single_auth_name.append("not_found")
            if row[2].value:
                single_auth_name.append(row[2].value)
            else:
                single_auth_name.append("not_found")
            if auth_type:
                single_auth_name.append(auth_type)
            else:
                single_auth_name.append("not_found")
            if row[4].value:
                single_auth_name.append(row[4].value)
            else:
                single_auth_name.append("not_found")
            auth_file_attr.append(single_auth_name)
            single_auth_name = []
        auth_lookup[reference] = auth_file_attr
    return auth_lookup


# Authority files process
def authority_files(row, arg, auth_lookup, E, shelfmark_modified, row_num):
    '''Takes the dictionry created in auth_lookup and separates each attribute
        to create the full authority file row.'''
    full_text = []
    if row[arg].value:
        lines = [""]
        if str(row[arg].value).find("|") != -1 :
            lines = row[arg].value.split("|")
        else:
            lines[0] = row[arg].value
        for line in lines:
            element_dict = {rel_persons_clmn: E.persname,
                            rel_fams_clmn: E.famname,
                            rel_corp_bds_clmn: E.corpname,
                            rel_places_clmn: E.geogname,
                            rel_subject_clmn: E.subject}
            current_auth = auth_lookup.get(int(line))
            for par_auth in current_auth:
                if arg == rel_persons_clmn or arg == rel_fams_clmn or arg == rel_corp_bds_clmn:
                    if arg == rel_persons_clmn and par_auth[0].find(",") != -1:
                        real_name = par_auth[0].split(",")
                        real_name = reversed(real_name)
                        par_auth[0] = " ".join(real_name)
                    text = element_dict[arg](par_auth[0],
                        {"authfilenumber": par_auth[1]},
                        {"role": par_auth[2]},
                        {"source": "IAMS"},
                        {"altrender": par_auth[3]},
                        tid(row, arg, shelfmark_modified, row_num))
                elif arg == rel_subject_clmn:
                    text = element_dict[arg](par_auth[0],
                        {"authfilenumber": par_auth[1]},
                        {"altrender": par_auth[3]},
                        {"source": "IAMS"},
                        tid(row, arg, shelfmark_modified, row_num))
                else:
                    text = element_dict[arg](par_auth[0],
                        {"authfilenumber": par_auth[1]},
                        {"role": par_auth[2]},
                        {"source": "IAMS"},
                        tid(row, arg, shelfmark_modified, row_num))
                full_text.append(text)
        return full_text
    else:
        return ""
        

# IAMS template validation definition
def template_verification(ws, sh_complete_label, sh_furth_steps_label):
    '''Checks the well-formedness of the IAMS template'''
    current_template_order = get_header(ws)
    approved_order = ['Repository','Collection Area','Project / Collection','Level','Reference','Former external reference','Title','Date range','Start date','End date','Era','Calendar','Extent','Scope and content','Physical characteristics','Access conditions','Immediate source of acquisition','Custodial history','Administrative context','Arrangement','Related material','Finding aids','Originals information','Copies information','Publication note','Exhibition','Related archival descriptions','Language of material','Language codes of material','Scripts of material','Script codes of material','Language of description','Language code of description','Script of description','Script code of description','Related persons','Related families','Related corporate bodies','Related places','Related subjects','Decimal Latitude','Decimal Longitude','Decimal Co-ordinates','Scale','Scale Designator','Projection','Orientation','Legal status','Level of detail','Visibility','Logical type','Logical label','Page sequence range','Page label range','Material type','Item type','ARK ID','IAMS ID']
    print(current_template_order == approved_order)
    if current_template_order == approved_order:
        validation_check = 0
        row_num = 0
        for row in ws.iter_rows(min_row=5):
            row_num += 1
            if len(row) == 58:
                if row[repository_clmn].value:
                    if row[coll_area_clmn].value:
                        if row[collection_clmn].value:
                            if row[level_clmn].value:
                                if row[reference_clmn].value:
                                    if row[title_clmn].value:
                                        if row[date_rng_clmn].value:
                                            if row[era_clmn].value:
                                                if row[calendar_clmn].value:
                                                    if row[access_cond_clmn].value:
                                                        if row[mat_language_clmn].value:
                                                            if row[mat_langcode_clmn].value:
                                                                if row[mat_script_clmn].value:
                                                                    if row[mat_scriptcode_clmn].value:
                                                                        if row[descr_lang_clmn].value:
                                                                            if row[mat_type_clmn].value:
                                                                                if row[scope_content_clmn].value:
                                                                                    if row[scope_content_clmn].value.find("<p>")!= -1:
                                                                                        if len(row[scope_content_clmn].value.split("<list>")) == len(row[scope_content_clmn].value.split("<p><list>")):
                                                                                            sh_complete_label.configure(text="Unrecognised Error", bg="#cc0000", fg="white")
                                                                                            validation_check += 1
                                                                                        else:
                                                                                            sh_furth_steps_label.config(text="Lists are not encased in <p><list> nodes" , fg="black")
                                                                                    else:
                                                                                        sh_furth_steps_label.config(text="Missing paragraph structure in free text" , fg="black")
                                                                                else:
                                                                                    validation_check += 1
                                                                            else:
                                                                                sh_furth_steps_label.config(text="Missing material type" , fg="black")
                                                                        else:
                                                                            sh_furth_steps_label.config(text="Missing description language" , fg="black")
                                                                    else:
                                                                        sh_furth_steps_label.config(text="Missing material script code" , fg="black")
                                                                else:
                                                                    sh_furth_steps_label.config(text="Missing material script" , fg="black")
                                                            else:
                                                                sh_furth_steps_label.config(text="Missing material language code" , fg="black")
                                                        else:
                                                            sh_furth_steps_label.config(text="Missing material language" , fg="black")
                                                    else:
                                                        sh_furth_steps_label.config(text="Missing access conditions" , fg="black")
                                                else:
                                                    sh_furth_steps_label.config(text="Missing calendar" , fg="black")
                                            else:
                                                sh_furth_steps_label.config(text="Missing era" , fg="black")
                                        else:
                                            sh_furth_steps_label.config(text="Missing date range" , fg="black")
                                    else:
                                        sh_furth_steps_label.config(text="Missing title" , fg="black")
                                else:
                                    sh_furth_steps_label.config(text="Missing shelfmark reference" , fg="black")
                            else:
                                sh_furth_steps_label.config(text="Missing record level" , fg="black")
                        else:
                            sh_furth_steps_label.config(text="Missing collection field" , fg="black")
                    else:
                        sh_furth_steps_label.config(text="Missing collection area field" , fg="black")
                else:
                    sh_furth_steps_label.config(text="Missing repository field" , fg="black")
            else:
                sh_furth_steps_label.config(text="Not enough fields in template" , fg="black")
        if validation_check == row_num:
            validated = True
        else:
            validated = False
    else:
        sh_furth_steps_label.config(text="Incorrect order of fields in IAMS template" , fg="black")
        validated = False
    return validated


# Full Gather process
def QatarGather(IAMS_filename, Auth_filename, end_directory):
    '''The main Qatar Gather code. This creates the full XML'''
    auth_file_wb = load_workbook(Auth_filename, read_only=True)
    auth_ws = auth_file_wb.active
    
    wb = load_workbook(IAMS_filename, read_only=True, data_only = True)
    shelfmarks = wb.sheetnames
    shm_num = 0

    for shelfmark_modified in shelfmarks:
        if shelfmark_modified != "Template" and shelfmark_modified != "Shelfmark Arks and IAMS ID" and shelfmark_modified != "Data Validation":
            shm_num += 1
            rec_num = 1
            row_num = 0
            wordcount = 0
            ws = wb[shelfmark_modified]
            header_row = get_header(ws)
            E = ElementMaker(namespace="urn:isbn:1-931666-22-9",
                             nsmap={'ead': "urn:isbn:1-931666-22-9",
                                    'xlink': "http://www.w3.org/1999/xlink",
                                    'xsi':
                                        "http://www.w3.org/2001/XMLSchema-instance"
                                    })
            full_ead = E.ead()

            sh_label = tk.Label(master=run_frame, text=shelfmark_modified)
            sh_label.grid(row=1+shm_num, column=0, padx=5, pady=5, sticky="nsew")
            sh_verif_lbl = tk.Label(master=run_frame)
            sh_verif_lbl.grid(row=1+shm_num, column=1, padx=5, pady=5,
                            sticky="nsew")
            sh_complete_label = tk.Label(master=run_frame)
            sh_complete_label.grid(row=1+shm_num, column=2, padx=5, pady=5,
                                sticky="nsew")
            sh_furth_steps_label = tk.Label(master=run_frame)
            sh_furth_steps_label.grid(row=1+shm_num, column=3, padx=5, pady=5,
                                sticky="nsew")
            complete_label = tk.Label(master=app, font=("calibri", 14, "bold"),
                                    anchor="e")
            complete_label.grid(row=5, column=0, padx=5, pady=5, sticky="nsew")
            sh_wordcount = tk.Label(master=run_frame)
            sh_wordcount.grid(row=1+shm_num, column=4, padx=5, pady=5,
                            sticky="nsew")
            if template_verification(ws, sh_complete_label, sh_furth_steps_label) is True:
                sh_verif_lbl.configure(text="Verified", font=("calibri", 10, "bold"), fg="#3b851a")
                if auth_validation(auth_ws) is True:
                    auth_lookup = gen_auth_lookup(auth_ws)

            # This part creates the tree for each child shelfmark.
                    for row in ws.iter_rows(min_row=5, values_only=False):
                        print(type(row))
                        ead = E.ead()
                        comment = Comment(
                            f"New record starts here {row[reference_clmn].value}")
                        full_ead.append(comment)
                        wordcount += current_wordcount(row)
                        shelfmark = str(row[reference_clmn].value)

                        # header
                        eadheader = E.eadheader(start_record(str(rec_num)))
                        ead.append(eadheader)

                        eadid = E.eadid(str(shelfmark),
                                        tid(row, reference_clmn,
                                            shelfmark_modified, row_num))
                        eadheader.append(eadid)
                        row_num += 1

                        filedesc = E.filedesc()  # wrapper node, should not have info
                        eadheader.append(filedesc)

                        titlestmt = E.titlestmt()  # wrapper node, should not have info
                        filedesc.append(titlestmt)

                        titleproper = E.titleproper()  # not used in IAMS material
                        titlestmt.append(titleproper)

                        profiledesc = E.profiledesc()  # wrapper node,  no info
                        eadheader.append(profiledesc)

                        creation = E.creation()  # not used in Qatar(?)
                        profiledesc.append(creation)

                        date_exported = E.date(str(datetime.now()
                                                .strftime("%Y-%m-%dT%H:%M:%S")),
                                            {"type": "exported"},
                                            tid(row, reference_clmn,
                                                shelfmark_modified, row_num))
                        creation.append(date_exported)

                        date_modified = E.date(str(wb.properties.modified.strftime(
                            "%Y-%m-%dT%H:%M:%S")), {"type": "modified"},
                            tid(row, reference_clmn, shelfmark_modified, row_num))
                        creation.append(date_modified)

                        langusage = E.langusage()  # not used in IAMS material
                        profiledesc.append(langusage)

                        # this is language of the description
                        language = E.language(content(row, descr_lang_clmn),
                                            labels(row, descr_langcode_clmn,
                                                    "langcode"),
                                            labels(row, descr_scriptcode_clmn,
                                                    "scriptcode"),
                                            tid(row, descr_lang_clmn,
                                                shelfmark_modified, row_num))
                        langusage.append(language)

                        # archdesc
                        archdesc = E.archdesc(labels(row, level_clmn, "level"))
                        ead.append(archdesc)

                        did = E.did()  # wrapper node, should not have info
                        archdesc.append(did)

                        # British Library: Indian Office Records
                        repository = E.repository(
                            row[repository_clmn].value + ": " + row[coll_area_clmn
                                                                    ].value,
                            tid(row, repository_clmn, shelfmark_modified, row_num))
                        did.append(repository)

                        unitid = E.unitid(shelfmark,
                                        labels(row, iams_id_clmn, "label"),
                                        labels(row, ark_id_clmn, "identifier"),
                                        tid(row, reference_clmn, shelfmark_modified,
                                            row_num))
                        # These are the IAMS identifiers (ark and number)
                        did.append(unitid)

                        # this will say "title"
                        unittitle = E.unittitle(header_label(header_row, title_clmn,
                                                            "label"))
                        did.append(unittitle)

                        # Item title
                        text_title = title_content(row, title_clmn, E, shelfmark_modified, row_num)
                        unittitle.append(text_title)

                        if row[ext_ref_clmn].value:
                            unittitle = E.unittitle(content(
                                row, ext_ref_clmn), header_label(header_row,
                                                                ext_ref_clmn,
                                                                "label"),
                                                    tid(row, ext_ref_clmn,
                                                        shelfmark_modified, row_num))
                        else:
                            unittitle = E.unittitle(
                                {"label": "Former external reference"})
                        did.append(unittitle)  # Former external reference

                        unittitle = E.unittitle({"label": "Former internal reference"})
                        did.append(unittitle)  # Former internal reference (not used)

                        unitdate = E.unitdate(content(row, date_rng_clmn), {
                            "datechar": "Creation"},
                            labels(row, calendar_clmn, "calendar"),
                            labels(row, era_clmn, "era"),
                            date_normal(row, date_rng_clmn),
                            tid(row, date_rng_clmn, shelfmark_modified, row_num))
                        did.append(unitdate)  # Date of the material

                        # This allows for multiple langs and langcodes separated by |
                        # (mat_language_clmn, mat_langcode_clmn) for language
                        # (mat_script_clmn, mat_scriptcode_clmn) for script
                        code_label_index = 0
                        for r in ((mat_language_clmn, mat_langcode_clmn),
                                (mat_script_clmn, mat_scriptcode_clmn)):
                            langmaterial = E.langmaterial()
                            did.append(langmaterial)
                            languages = row[r[0]].value.split("|")
                            lang_codes = row[r[1]].value.split("|")
                            code_labels = ["langcode", "scriptcode"]
                            for lang, c in zip(languages, lang_codes):
                                language = E.language(
                                    lang, {code_labels[code_label_index]: c},
                                    tid(row, r[0], shelfmark_modified, row_num))
                                langmaterial.append(language)
                            code_label_index += 1

                        physdesc = E.physdesc()  # wrapper node, should not have info
                        did.append(physdesc)

                        extent = E.extent(content(row, extent_clmn),
                                        tid(row, extent_clmn, shelfmark_modified, row_num))
                        physdesc.append(extent)
                # Map details generated here
                        if row[mat_type_clmn].value == 'Maps and Plans':
                            if row[scope_content_clmn].value:
                                materialspec = E.materialspec(content(row, scale_clmn),
                                                            {'type': "scale"},
                                                            tid(row, scale_clmn,
                                                                shelfmark_modified,
                                                                row_num))
                                did.append(materialspec)
                                materialspec = E.materialspec(
                                    content(row, scale_des_clmn),
                                    {'type': "scale designator"},
                                    tid(row, scale_des_clmn, shelfmark_modified,
                                        row_num))
                                did.append(materialspec)
                                materialspec = E.materialspec(
                                    content(row, dec_coords_clmn),
                                    {'type': "coordinates"}, {'label': "decimal"},
                                    tid(row, dec_coords_clmn, shelfmark_modified,
                                        row_num))
                                did.append(materialspec)
                                materialspec = E.materialspec(
                                    content(row, orientation_clmn),
                                    {'type': "orientation"}, tid(row, orientation_clmn,
                                                                shelfmark_modified,
                                                                row_num))
                                did.append(materialspec)


                        text = row[access_cond_clmn].value.split("</list></p>")
                        accessrestrict = E.accessrestrict()
                        for i in text:
                            if i:
                                cnt = pcontent(i, row, access_cond_clmn, E, shelfmark_modified, row_num)
                                for l in cnt:
                                    accessrestrict.append(l)
                        archdesc.append(accessrestrict)

                        accessrestrict = E.accessrestrict()
                        # This second accessrestrict is a wrapper node
                        archdesc.append(accessrestrict)

                        legalstatus = E.legalstatus(content(row, legal_sts_clmn),
                                                    tid(row, legal_sts_clmn,
                                                        shelfmark_modified, row_num))
                        accessrestrict.append(legalstatus)

                        accruals = E.accruals()  # Empty node
                        p = E.p()
                        accruals.append(p)
                        archdesc.append(accruals)

                        bioghist = E.bioghist()
                        if row[admin_context_clmn].value:
                            text = row[admin_context_clmn].value.split("</list></p>")
                            for i in text:
                                if i:
                                    cnt = pcontent(i, row, admin_context_clmn, E, shelfmark_modified, row_num)
                                    for l in cnt:
                                        bioghist.append(l)
                            archdesc.append(bioghist)
                        else:
                            p = E.p()
                            bioghist.append(p)
                            archdesc.append(bioghist)

                        appraisal = E.appraisal()  # Empty node
                        p = E.p()
                        appraisal.append(p)
                        archdesc.append(appraisal)

                        arrangement = E.arrangement()
                        if row[arrangement_clmn].value:
                            text = row[arrangement_clmn].value.split("</list></p>")
                            for i in text:
                                if i:
                                    cnt = pcontent(i, row, arrangement_clmn, E, shelfmark_modified, row_num)
                                    for l in cnt:
                                        arrangement.append(l)
                            archdesc.append(arrangement)
                        else:
                            p = E.p()
                            arrangement.append(p)
                            archdesc.append(arrangement)

                        # Adds custodial history if necessary
                        if row[cust_hist_clmn].value:
                            text = row[cust_hist_clmn].value.split("</list></p>")
                            custodhist = E.custodhist()
                            for i in text:
                                if i:
                                    print("line =", i)
                                    cnt = pcontent(i, row, cust_hist_clmn, E, shelfmark_modified, row_num)
                                    for l in cnt:
                                        custodhist.append(l)
                            archdesc.append(custodhist)
                        
                        # Adds finding aids if necessary
                        if row[find_aids_clmn].value:
                            text = row[find_aids_clmn].value.split("</list></p>")
                            otherfindaid = E.otherfindaid()
                            for i in text:
                                if i:
                                    cnt = pcontent(i, row, find_aids_clmn, E, shelfmark_modified, row_num)
                                    for l in cnt:
                                        otherfindaid.append(l)
                            archdesc.append(otherfindaid)
                        
                        # Adds pubblication notes if necessary
                        if row[pub_notes_clmn].value:
                            text = row[pub_notes_clmn].value.split("</list></p>")
                            bibliography = E.bibliography()
                            for i in text:
                                if i:
                                    cnt = pcontent(i, row, pub_notes_clmn, E, shelfmark_modified, row_num)
                                    for l in cnt:
                                        bibliography.append(l)
                            archdesc.append(bibliography)

                        # Adds acquisition notes if necessary
                        if row[imm_acq_column].value:
                            text = row[imm_acq_column].value.split("</list></p>")
                            acqinfo = E.acqinfo()
                            for i in text:
                                if i:
                                    cnt = pcontent(i, row, imm_acq_column, E, shelfmark_modified, row_num)
                                    for l in cnt:
                                        acqinfo.append(l)
                            archdesc.append(acqinfo)

                        # This allows to skip the node if item is part of a bigger volume
                        if row_num == 1 or row[mat_type_clmn].value != "Archives and Manuscripts":
                            phystech = E.phystech()
                            if row[phys_char_clmn].value:
                                text = row[phys_char_clmn].value.split("</list></p>")
                                for i in text:
                                    if i:
                                        cnt = pcontent(i, row, phys_char_clmn, E, shelfmark_modified, row_num)
                                        for l in cnt:
                                            phystech.append(l)
                                archdesc.append(phystech)
                            else:
                                p = E.p()
                                phystech.append(p)
                                archdesc.append(phystech)

                        # Adds Scope and Content field
                        scopecontent = E.scopecontent()
                        if row[scope_content_clmn].value:
                            text = row[scope_content_clmn].value.split("</list></p>")
                            for i in text:
                                if i:
                                    cnt = pcontent(i, row, scope_content_clmn, E, shelfmark_modified, row_num)
                                    for l in cnt:
                                        scopecontent.append(l)
                            archdesc.append(scopecontent)
                        else:
                            p = E.p()
                            scopecontent.append(p)
                            archdesc.append(scopecontent)

                        userestrict = E.userestrict()  # Empty node
                        p = E.p()
                        userestrict.append(p)
                        archdesc.append(userestrict)

                        odd = E.odd()  # Empty node
                        p = E.p()
                        odd.append(p)
                        archdesc.append(odd)

                        controlaccess = E.controlaccess()
                        genreform = E.genreform(
                            content(row, mat_type_clmn), {"source": "IAMS"},
                            tid(row, mat_type_clmn, shelfmark_modified, row_num))
                        controlaccess.append(genreform)

                        # Authority files processing starts here:
                        auth_lookup = gen_auth_lookup(auth_ws)
                        for arg in range(rel_persons_clmn, rel_subject_clmn+1, 1):
                            for af in authority_files(row, arg, auth_lookup, E,
                                                    shelfmark_modified, row_num):
                                controlaccess.append(af)
                        archdesc.append(controlaccess)
                        # End of authority files

                        project = []
                        project = row[collection_clmn].value.split("|")
                        for i in project:
                            note = E.note({"type": "project/collection"})
                            p = E.p(i, tid(row, mat_type_clmn, shelfmark_modified, row_num))
                            note.append(p)
                            controlaccess.append(note)

                        rec_num += 1

            # This puts together two parts of each child's tree (header+description)
            # This will append as many children as there are in the Excel tab
                        full_ead.append(eadheader)
                        full_ead.append(archdesc)

            # This part writes out the XML file
                    file_name = "Translation_English_"+shelfmark_modified+"_"+str(datetime.now().strftime("%Y%m%d_%H%M"))+'.xml'
                    with open(end_directory+"/"+ file_name, 'wb') as f:
                        f.write(etree.tostring(
                            full_ead, encoding="utf-8", xml_declaration=True,
                            pretty_print=True))
                        

                    sh_complete_label.config(text="Complete", fg="green", bg="#F0F0F0", font=("calibri",10,"bold"))
                    sh_wordcount.config(text=wordcount)
                else:
                    sh_verif_lbl.configure(text="Authority Files sheet not recognised", bg="#cc0000",
                                        fg="white")
                    sh_complete_label.configure(text="Unable to complete", fg="#cc0000", bg ="#F0F0F0")
                    sh_furth_steps_label.config(text="Check the Authority Files sheet" , fg="black")

            else:
                sh_verif_lbl.configure(text="IAMS Template recognised", bg="#cc0000",
                                        fg="white")
                sh_complete_label.configure(text="Unable to complete", fg="#cc0000", bg ="#F0F0F0")
                # sh_furth_steps_label.config(text="Check the IAMS Template sheet" , fg="black")
        
    wb.close()
    auth_file_wb.close()
    status = "Process complete"
    complete_label.config(text=status)


# App Editor Definitions
def UploadIAMS(event=None):
    global IAMS_filename
    IAMS_filename = filedialog.askopenfilename(
        filetypes=(("Excel files", "*.xlsx"), ("Any file", "*")))
    IAMS_filename_label.delete(0, END)
    IAMS_filename_label.insert(0, IAMS_filename)


def UploadAuth(event=None):
    global auth_filename
    auth_filename = filedialog.askopenfilename(
        filetypes=(("Excel files", "*.xlsx"), ("Any file", "*")))
    auth_filename_label.delete(0, END)
    auth_filename_label.insert(0, auth_filename)


def askDirectory(event=None):
    global end_directory
    end_directory = filedialog.askdirectory()
    end_directory_label.delete(0, END)
    end_directory_label.insert(0, end_directory)


app = tk.Tk()
app.title("Gather Renewed")
app.option_add("*font", "calibri 10")
app.minsize(630, 220)

title_lbl = tk.Label(master=app, text="Gather Renewed", bg="#F0F0F0", fg="black", anchor="w", font=("calibri", 18, "bold"))
title_lbl.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

frame = tk.Frame(app)
frame.grid(row=1, column=0, sticky="nsew")
canvas = tk.Canvas(frame)
scrollbar = tk.Scrollbar(frame, orient="vertical", command=canvas.yview)
canvas.configure(yscrollcommand=scrollbar.set)
content_frame = tk.Frame(canvas)
content_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

#title_lbl = tk.Label(master=content_frame, text="Gather Renewed", bg="#F0F0F0",
                     #fg="black", anchor="w", font=("calibri", 18, "bold"))
instr_lbl = tk.Label(master=content_frame,
                     text="Please select the IAMS template to gather, the relevant Authority File sheet and a destination folder for the created files. Once ready, click 'Run'.",
                     anchor="w", wraplength= 600, justify = LEFT)


# IAMS Fields
selection_frm = tk.LabelFrame(master=content_frame, text="Select Files", bg="#F0F0F0",
                              fg="black")
IAMS_label = tk.Label(master=selection_frm, text="Import IAMS Template",
                      bg="#F0F0F0", fg="black")
IAMS_button = tk.Button(selection_frm, text='Open', bg="#0b5394", fg="white",
                        command=UploadIAMS)
IAMS_filename_label = tk.Entry(master=selection_frm, width = 50)
# AuthFiles Fields
auth_file_label = tk.Label(master=selection_frm,
                           text="Import Authorities Spreadsheet", bg="#F0F0F0",
                           fg="black")
auth_file_button = tk.Button(selection_frm, text='Open', bg="#0b5394",
                             fg="white", command=UploadAuth)
auth_filename_label = tk.Entry(master=selection_frm)
# Destination Fields
dir_label = tk.Label(master=selection_frm,
                     text="Select directory to save files to", bg="#F0F0F0",
                     fg="black")
dir_button = tk.Button(selection_frm, text='Open', bg="#0b5394", fg="white",
                       command=askDirectory)
end_directory_label = tk.Entry(master=selection_frm)
# Run button
run_button = tk.Button(master=content_frame, text="Run", bg="green", fg="white",
                       command=lambda: QatarGather(IAMS_filename,
                                                   auth_filename,
                                                   end_directory))

# Running frame

run_frame = tk.LabelFrame(master=content_frame, text="Running", bg="#F0F0F0", fg="black")

run_shmark = tk.Label(master=run_frame, text="Shelfmark", bg="#F0F0F0",
                      fg="black")
run_verification = tk.Label(master=run_frame, text="IAMS Template Validation",
                            bg="#F0F0F0", fg="black")
run_status = tk.Label(master=run_frame, text="Status", bg="#F0F0F0",
                      fg="black")
run_further_steps = tk.Label(master=run_frame, text="Further Steps", bg="#F0F0F0",
                      fg="black")
run_wordcount = tk.Label(master=run_frame, text="Wordcount", bg="#F0F0F0",
                         fg="black")

title_lbl.grid(column=0, row=0, sticky="nsew", padx=5)
instr_lbl.grid(column=0, row=1, sticky="nsew", padx=5)
selection_frm.grid(column=0, row=2, sticky="nsew", padx=5)

IAMS_label.grid(column=0, row=0, columnspan=1, sticky="nsew", padx=5, pady=5)
IAMS_button.grid(column=1, row=0, columnspan=1, sticky="nsew", padx=5, pady=5)
IAMS_filename_label.grid(column=2, row=0, columnspan=3, sticky="nsew", padx=5,
                         pady=5)

auth_file_label.grid(column=0, row=1, columnspan=1, sticky="nsew", padx=5,
                     pady=5)
auth_file_button.grid(column=1, row=1, columnspan=1, sticky="nsew", padx=5,
                      pady=5)
auth_filename_label.grid(column=2, row=1, columnspan=3, sticky="nsew", padx=5,
                         pady=5)

dir_label.grid(column=0, row=2, columnspan=1, sticky="nsew", padx=5, pady=5)
dir_button.grid(column=1, row=2, columnspan=1, sticky="nsew", padx=5, pady=5)
end_directory_label.grid(column=2, row=2, columnspan=3, sticky="nsew", padx=5,
                         pady=5)

run_button.grid(column=0, row=3, sticky="nsew", padx=5, pady=5)

run_frame.grid(column=0, row=4, sticky="nsew", padx=5, pady=5)
run_shmark.grid(column=0, row=0, sticky="nsew", padx=20, pady=5)
run_verification.grid(column=1, row=0, sticky="nsew", padx=20, pady=5)
run_status.grid(column=2, row=0, sticky="nsew", padx=20, pady=5)
run_further_steps.grid(column=3, row=0, sticky="nsew", padx=20, pady=5)
run_wordcount.grid(column=4, row=0, sticky="nsew", padx=20, pady=5)

app.columnconfigure(0, weight=1)
frame.columnconfigure(0, weight=1)
frame.rowconfigure(0, weight=1)


canvas.create_window((0, 0), window=content_frame, anchor="nw")
canvas.grid(row=0, column=0, sticky="nsew")
scrollbar.grid(row=0, column=1, sticky="ns")

def _on_mousewheel(event):
   canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

canvas.bind_all("<MouseWheel>", _on_mousewheel)

app.mainloop()
