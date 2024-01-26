# -*- coding: utf-8 -*-
"""
Created on Thu Jan 25 17:09:02 2024

@author: amilighe

The code should be stored in the same folder as the gather sheet. The output
will also be in the same folder.

Notes on the creation of the template:
    - Authority files should be in teh format:
            Name=role=altrender|Name=role=altreder
      with no extra spaces.
      The authority file sheet's name is hard-coded in the code.
    - The authority file sheet should be updated to match the authority files
      in each item to Gather. If a new sheet is created, update the name of the
      sheet in line 144.
    - The IAMS template sheet should have one tab for each item
      and one row for each child.
    - When inputting the IAMS sheet's name.
      Omit the .xlsx as it will be added in the code
    - Note that Authority file processing takes a longer time as it needs to
      reference a high volume of items.
    - The code only has very basic error control. Always check the output.
"""
from datetime import datetime

from lxml import etree
from lxml.builder import ElementMaker
from lxml.etree import Comment
from openpyxl import load_workbook


# These are the column number for the fields used
# If the template changes, change the column numbers here
# IAMS template:
repository_clmn = 0
coll_area_clmn = 1
collection_clmn = 2
level_clmn = 4
reference_clmn = 5
ext_ref_clmn = 6
title_clmn = 7
date_rng_clmn = 8
era_clmn = 9
calendar_clmn = 10
extent_clmn = 11
scope_content_clmn = 12
phys_char_clmn = 13
access_cond_clmn = 14
arrangement_clmn = 15
mat_language_clmn = 23
mat_langcode_clmn = 24
mat_script_clmn = 25
mat_scriptcode_clmn = 26
descr_lang_clmn = 27
descr_langcode_clmn = 28
descr_script_clmn = 29
descr_scriptcode_clmn = 30
rel_persons_clmn = 31
rel_fams_clmn = 32
rel_corp_bds_clmn = 33
rel_places_clmn = 34
rel_subject_clmn = 35
legal_sts_clmn = 43
mat_type_clmn = 50
# Authority files document
auth_name_clmn = 36
auth_ark_id_clmn = 9


# Definitions used to create the nodes:

def get_header(ws):
    header = []
    for cell in ws[1]:
        header.append(cell.value)
    return header


def start_record(rec_num):
    return {"StartRecord": rec_num}


def tid(row, arg):
    # the tid_num should grow within the xml across child records
    global tid_num
    if row[arg].value:
        tid_full = shelfmark_modified+"_"+str(tid_num)
        tid_num = tid_num+1
        return {"tid": tid_full}
    else:
        return {}


def content(row, arg):
    if row[arg].value:
        return str(row[arg].value)
    else:
        return {}


def labels(row, arg, label):
    content = str(row[arg].value)
    return {label: content}


def header_label(header_row, arg, label):
    label_title = header_row[arg]
    return {label: label_title}


def date_normal(row, arg):
    full_date = str(row[arg].value)
    if "-" in full_date:
        index = full_date.rfind("-")
        start_date = full_date[index-4:index]
        end_date = full_date[-4:]
        date = start_date+"/"+end_date
    else:
        date = full_date[-4:]+"/"+full_date[-4:]
    return {"normal": date}


def pcontent(row, arg):
    content = []
    if row[arg].value:
        lines = row[arg].value.replace("</p>", "").split("<p>")
        for line in lines:
            if line:
                p = E.p(line, tid(row, arg))
                content.append(p)
    else:
        p = E.p()
        content.append(p)
    return content


# Authority Files processing definitions
def gen_auth_lookup(auth_ws, auth_name_clmn, auth_ark_id_clmn):
    auth_lookup = {}
    for row_num, row in enumerate(auth_ws.iter_rows()):
        auth_name = str(row[auth_name_clmn].value).strip().lower()
        if row[auth_ark_id_clmn].value:
            ark_id = row[auth_ark_id_clmn].value
        else:
            ark_id = "not_allocated"
        auth_lookup[auth_name] = ark_id
    return auth_lookup


def auth_dets(arg, label):
    if arg == "not_allocated":
        return {}
    else:
        return {label: arg}


def authority_files(row, arg, auth_lookup):
    if row[arg].value:
        lines = row[arg].value.split("|")
        full_text = []
        for line in lines:
            attributes = line.split("=")
            subject = attributes[0]
            role_type = "not_allocated"
            altrender_type = "not_allocated"
            if len(attributes) > 1 and attributes[1]:
                role_type = attributes[1]
            if len(attributes) > 2:
                altrender_type = attributes[2]
            element_dict = {rel_persons_clmn: E.persname,
                            rel_fams_clmn: E.famname,
                            rel_corp_bds_clmn: E.corpname,
                            rel_places_clmn: E.geogname,
                            rel_subject_clmn: E.subject}
            text = element_dict[arg](
                subject,
                {"authfilenumber":
                 auth_lookup.get(subject.strip().lower(), "not_found")},
                auth_dets(role_type, "role"),
                {"source": "IAMS"},
                auth_dets(altrender_type, "altrender"),
                tid(row, arg))

            full_text.append(text)
        return full_text
    else:
        return ""


# The actual code starts here.
# This part defines where the authority files details are held.
auth_file_name = 'TB_Authorities.xlsx'
auth_file_wb = load_workbook(auth_file_name, read_only=True)
auth_ws = auth_file_wb["in"]
auth_lookup = gen_auth_lookup(auth_ws, auth_name_clmn, auth_ark_id_clmn)

# This selects the spreadsheet to gather
# One tab each shelfmark to gather.
wb_input = input(
    'Please write the name of the spreadsheet to Gather. Omit ".xlsx": ')
wb_name = wb_input + '.xlsx'
wb = load_workbook(wb_name, read_only=True)
shelfmarks = wb.sheetnames
for shelfmark_modified in shelfmarks:
    rec_num = 1
    tid_num = 1
    row_num = 1
    ws = wb[shelfmark_modified]
    header_row = get_header(ws)

    E = ElementMaker(namespace="urn:isbn:1-931666-22-9",
                     nsmap={'ead': "urn:isbn:1-931666-22-9",
                            'xlink': "http://www.w3.org/1999/xlink",
                            'xsi': "http://www.w3.org/2001/XMLSchema-instance"
                            })
    full_ead = E.ead()

# This part creates the tree for each child shelfmark.

    for row in ws.iter_rows(min_row=2, values_only=False):
        ead = E.ead()
        comment = Comment(
            f"New record starts here {row[reference_clmn].value}")
        full_ead.append(comment)
        shelfmark = str(row[reference_clmn].value)
        print(shelfmark)

        # header
        eadheader = E.eadheader(start_record(str(rec_num)))
        ead.append(eadheader)

        eadid = E.eadid(str(shelfmark), tid(row, reference_clmn))
        eadheader.append(eadid)

        filedesc = E.filedesc()  # wrapper node, should not have info
        eadheader.append(filedesc)

        titlestmt = E.titlestmt()  # wrapper node, should not have info
        filedesc.append(titlestmt)

        titleproper = E.titleproper()  # not used in IAMS material
        titlestmt.append(titleproper)

        profiledesc = E.profiledesc()  # wrapper node, should not have info
        eadheader.append(profiledesc)

        creation = E.creation()  # not used in Qatar(?)
        profiledesc.append(creation)

        date_exported = E.date(str(datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
                                   ), {"type": "exported"},
                               tid(row, reference_clmn))
        creation.append(date_exported)

        date_modified = E.date(str(wb.properties.modified.strftime(
            "%Y-%m-%dT%H:%M:%S")), {"type": "modified"},
            tid(row, reference_clmn))
        creation.append(date_modified)

        langusage = E.langusage()  # not used in IAMS material
        profiledesc.append(langusage)

        # this is language of the description
        language = E.language(content(row, descr_lang_clmn),
                              labels(row, descr_langcode_clmn, "langcode"),
                              labels(row, descr_scriptcode_clmn, "scriptcode"),
                              tid(row, descr_lang_clmn))
        langusage.append(language)

        # archdesc
        archdesc = E.archdesc(labels(row, level_clmn, "level"))
        ead.append(archdesc)

        did = E.did()  # wrapper node, should not have info
        archdesc.append(did)

        # British Library: Indian Office Records
        repository = E.repository(
            row[repository_clmn].value + ": " + row[coll_area_clmn].value,
            tid(row, repository_clmn))
        did.append(repository)

        unitid = E.unitid(shelfmark, {"label": "IAMS_label_NA"},
                          {"identifier": "ark_identifier"},
                          tid(row, reference_clmn))
        # These are the IAMS identifiers (ark and number)
        did.append(unitid)

        # this will say "title"
        unittitle = E.unittitle(header_label(header_row, title_clmn, "label"))
        did.append(unittitle)

        # Item title
        title = E.title(content(row, title_clmn), tid(row, title_clmn))
        unittitle.append(title)

        if row[ext_ref_clmn].value:
            unittitle = E.unittitle(content(row, ext_ref_clmn), header_label(
                header_row, ext_ref_clmn, "label"), tid(row, ext_ref_clmn))
        else:
            unittitle = E.unittitle({"label": "Former external reference"})
        did.append(unittitle)  # Former external reference

        unittitle = E.unittitle({"label": "Former internal reference"})
        did.append(unittitle)  # Former internal reference (not used)

        unitdate = E.unitdate(content(row, date_rng_clmn), {
            "datechar": "Creation"}, labels(row, calendar_clmn, "calendar"),
            labels(row, era_clmn, "era"), date_normal(row, date_rng_clmn),
            tid(row, date_rng_clmn))
        did.append(unitdate)  # Date of the material

        # langmaterial = E.langmaterial()  # This is language
        # did.append(langmaterial)

        # This allows for multiple languages and language codes separated by |
        # (mat_language_clmn, mat_langcode_clmn) for language
        # (mat_script_clmn, mat_scriptcode_clmn) for script
        code_label_index = 0
        for r in ((mat_language_clmn, mat_langcode_clmn),
                  (mat_script_clmn, mat_scriptcode_clmn)):
            langmaterial = E.langmaterial()
            did.append(langmaterial)
            languages = row[r[0]].value.split("|")
            lang_codes = row[r[1]].value.split("|")
            code_labels = ["langcode", "scriptcode"]
            for lang, c in zip(languages, lang_codes):
                language = E.language(lang, {code_labels[code_label_index]: c},
                                      tid(row, r[0]))
                langmaterial.append(language)
            code_label_index += 1

        physdesc = E.physdesc()  # wrapper node, should not have info
        did.append(physdesc)

        extent = E.extent(content(row, extent_clmn), tid(row, extent_clmn))
        physdesc.append(extent)

        accessrestrict = E.accessrestrict()
        for p in pcontent(row, access_cond_clmn):
            accessrestrict.append(p)
        archdesc.append(accessrestrict)

        accessrestrict = E.accessrestrict()
        # This second accessrestrict is a wrapper node
        archdesc.append(accessrestrict)

        legalstatus = E.legalstatus(content(row, legal_sts_clmn),
                                    tid(row, legal_sts_clmn))
        accessrestrict.append(legalstatus)

        accruals = E.accruals()  # Empty node
        p = E.p()
        accruals.append(p)
        archdesc.append(accruals)

        bioghist = E.bioghist()  # Empty node
        p = E.p()
        bioghist.append(p)
        archdesc.append(bioghist)

        appraisal = E.appraisal()  # Empty node
        p = E.p()
        appraisal.append(p)
        archdesc.append(appraisal)

        arrangement = E.arrangement()
        for p in pcontent(row, arrangement_clmn):
            arrangement.append(p)
        archdesc.append(arrangement)

        phystech = E.phystech()
        for p in pcontent(row, phys_char_clmn):
            phystech.append(p)
        archdesc.append(phystech)

        # This section allows for bullet points in the scope and content
        scopecontent = E.scopecontent()
        lists = E.list()
        if row[scope_content_clmn].value:
            if row[scope_content_clmn].value.find("<list>") != -1:
                list_content = []
                paragraphs = row[scope_content_clmn].value.split("</p>")
                for chunk in paragraphs:
                    lines = chunk.split("</item>")
                    for line in lines:
                        line = line.strip()
                        if line:
                            tid_label = tid(row, scope_content_clmn)
                            if line.find('<emph render="italic">') != -1:
                                top = line.split('<emph render="italic">')[0]
                                emph_all = line.split(
                                    '<emph render="italic">')[1]
                                emph = emph_all.split('</emph>')[0]
                                bottom = line.split("</emph>")[1]
                                emph_tid = shelfmark_modified+"_"+str(tid_num)
                                tid_num += 1
                                line = top + '<ead:emph render="italic" tid="'
                                + emph_tid + '">' + emph + '</ead:emph>'
                                + bottom
                            line = line.replace("<list>", "").replace(
                                "</list>", "")
                            if line.startswith("<item>"):
                                line = line.replace("<item>", "")
                                list_content.append(line)
                                line_content = E.item(line, tid_label)
                                lists.append(line_content)
                                scopecontent.append(lists)
                            elif list_content == []:
                                line = line.replace("<p>", "")
                                top_p = E.p(line, tid_label)
                                scopecontent.append(top_p)
                            else:
                                line = line.replace("<p>", "")
                                bttm_p = E.p(line,
                                             tid_label)
                                scopecontent.append(bttm_p)
        else:
            for p in pcontent(row, scope_content_clmn):
                scopecontent.append(p)
        archdesc.append(scopecontent)

        userestrict = E.userestrict()  # Empty node
        p = E.p()
        userestrict.append(p)
        archdesc.append(userestrict)

        odd = E.odd()  # Empty node
        p = E.p()
        odd.append(p)
        archdesc.append(odd)

        controlaccess = E.controlaccess()
        genreform = E.genreform(
            content(row, mat_type_clmn), {"source": "IAMS"},
            tid(row, mat_type_clmn))
        controlaccess.append(genreform)

        # Authority files processing starts here:
        for arg in range(rel_persons_clmn, rel_subject_clmn+1, 1):
            for af in authority_files(row, arg, auth_lookup):
                controlaccess.append(af)
        archdesc.append(controlaccess)
        # End of authority files

        note = E.note({"type": "project/collection"})
        for p in pcontent(row, coll_area_clmn):
            note.append(p)
        controlaccess.append(note)

        note = E.note({"type": "project/collection"})
        for p in pcontent(row, collection_clmn):
            note.append(p)
        controlaccess.append(note)

        rec_num += 1
        row_num += 1

# This puts together the two parts of each child's tree (header+description)
# This will append as many children as there are in the Excel tab
        full_ead.append(eadheader)
        full_ead.append(archdesc)

# This part writes out the XML file
    with open(shelfmark_modified+"_"+str(datetime.now().strftime(
            "%Y%m%d_%H%M%S"))+'.xml', 'wb') as f:
        f.write(etree.tostring(full_ead, encoding="utf-8",
                               xml_declaration=True, pretty_print=True))

    print(shelfmark + ' complete \n')

wb.close()
auth_file_wb.close()
print('Gather complete!')
