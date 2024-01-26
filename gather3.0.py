# -*- coding: utf-8 -*-
"""
Created on Mon Jan  8 09:23:07 2024

@author: amilighe

The code shoudl be stored in the same folder as the gather sheet. The output
will also be in the same folder. 

Notes on the creation of the Template:
    - Authority files should be in teh format: Name>role>altrender|Name>role>altreder
      with no extra spaces. The authority file sheet's name is hard-coded in the code.
    - The authority file sheet should be updated to match the authority files 
      in each item to Gather. If a new sheet is created, update the name of the 
      sheet in line 202.
    - The IAMS template sheet should have one tab for each itema nd one row for each child.
    - When inputting the IAMS sheet, omit the .xlsx as it will be added in the code
    - Note that Authority file processing takes a longer time as it needs to 
      reference a high volume of items.
    - The code only has very basic error control. Always check the output.
"""

from lxml import etree
from lxml.builder import ElementMaker
from lxml.etree import Comment
from datetime import datetime
from openpyxl import load_workbook

#definitions of the node contents:
    
def get_header():
    header = []
    for cell in ws[1]:  
        header.append(cell.value)
    return header
def StartRecord (rec_num): 
    return {"StartRecord":rec_num}
def tid (row,arg):
    global tid_num
    if row[arg].value != None:
        fixed_shelfmark = row[5].value.replace("/","_").replace(" ","_").replace("-","_").replace(",","_")
        tid_full = fixed_shelfmark+"_"+str(tid_num)
        tid_num=tid_num+1
        return{"tid":tid_full}
    else:
        return""
def date_type(arg):
    return{"type":arg}
def langcode (row):
    lang_code = row[45].value
    return{"langcode":lang_code}
def scriptcode (row):
    script_code = row[47].value
    return{"scriptcode":script_code}
def level(row):
    item_level = row[4].value
    return{"level":item_level}
def IAMSlabel(row):
    return{"label":"IAMS_label_NA"}
def identifier(row):
    return{"identifier":"ark_identifier"}
def label(arg,header_row):
    label_title=header_row[arg]
    return{"label":label_title}
def datechar(row):
    return{"datechar":"Creation"}
def calendar(row):
    calendar_type = row[18].value
    return{"calendar":calendar_type}
def era(row):
    era_type = row[17].value
    return{"era":era_type}
def normal(row):
    normal_type = str(row[14].value)
    return{"normal":normal_type}
def date_in_full(row):
    if row[16].value:
        date_item = str(row[15].value)+"-"+str(row[16].value)
    else:
       date_item = str(row[15].value) 
    return date_item
def mat_langcode(c):
    #lang_code = row[41].value
    return{"langcode":c}
def mat_scriptcode(row):
    script_code = row[43].value
    return{"scriptcode":script_code}
def source(row):
    source = "IAMS"
    return {"source":source}
def authfilenumber(auth_row,auth_row_num):
    ark_id_auth = auth_row[18].value
    return {"authfilenumber":ark_id_auth}
def authfilenumberna(arg):
    return {"authfilenumber":arg}

def role (role_type):
    if role_type == "not_allocated":
        return{}
    if role_type == "":
        return {}
    else:
        return {"role":role_type}

def altrender (altrender_type):
    if altrender_type == "not_allocated":
        return{}
    else:
        return {"altrender":altrender_type}

def lang_content(l):
    if l:
        return l
    else:
        return""
        
def content(row,arg):
    if row[arg].value:
        return row[arg].value    
    else:
       return ""

def pcontent(row, arg):
    content = []   
    if row[arg].value:      
        lines = row[arg].value.split("\n")           
        for line in lines:
            p = E.p(line, tid(row,arg))   
            content.append(p)
    else:
        p = E.p("")
        content.append(p)
    return content

def search_auth_file(line, auth_ws):
    #print(f"Searching for: {line}")
    for row_num, row in enumerate(auth_ws.iter_rows()):
        cell = row[0]
        if str(cell.value).strip().lower() == str(line).strip().lower():
            #print (str(cell.value))
            return row_num+1
   # print("Authority file not found")
    return "not_found"
    

def authorityfiles(row,arg):
    if row[arg].value:
        lines = row[arg].value.split("|")
        #print(lines)
        full_text = []
        for line in lines:
           # print (line)
            
            attributes = line.split(">")
            subject = attributes[0]
            if len(attributes)>1:
                role_type = attributes[1]
            else:
                role_type = "not_allocated"
            if len(attributes)>2:    
                altrender_type=attributes[2]
            else:
                altrender_type="not_allocated"
            auth_row_num = search_auth_file(subject, auth_ws)
            if auth_row_num != "not_found":
                #print(f"Auth row num: {auth_row_num}")
                auth_row = auth_ws[auth_row_num]
                #print(auth_row[4].value)
                if auth_row[4].value == "Corporate Body":
                    text = E.corpname(subject,authfilenumber(auth_row,auth_row_num),role(role_type),source(auth_row),altrender(altrender_type),tid(row,arg))
                elif auth_row[4].value == "Person":
                    text = E.persname(subject,authfilenumber(auth_row,auth_row_num),role(role_type),source(auth_row),altrender(altrender_type),tid(row,arg))
                elif auth_row[4].value == "Family":
                    text = E.famname(subject,authfilenumber(auth_row,auth_row_num),role(role_type),source(auth_row),altrender(altrender_type),tid(row,arg))
                elif auth_row[4].value == "Place":
                    text = E.geogname(subject,authfilenumber(auth_row,auth_row_num),role(role_type),source(auth_row),altrender(altrender_type),tid(row,arg))
                elif auth_row[4].value == "Subject":
                    text = E.subject(subject,authfilenumber(auth_row,auth_row_num),role(role_type),source(auth_row),altrender(altrender_type),tid(row,arg))
                full_text.append(text)
            else: 
                if arg == 48:
                    text = E.persname(subject,authfilenumberna("not_found"),role(role_type),source(row),altrender(altrender_type),tid(row,arg))
                elif arg == 49:
                    text = E.famname(subject,authfilenumberna("not_found"),role(role_type),source(row),altrender(altrender_type),tid(row,arg))
                elif arg == 50:
                    text = E.corpname(subject,authfilenumberna("not_found"),role(role_type),source(row),altrender(altrender_type),tid(row,arg))
                elif arg == 51:
                    text = E.geogname(subject,authfilenumberna("not_found"),role(role_type),source(row),altrender(altrender_type),tid(row,arg))
                elif arg == 52:
                    text = E.subject(subject,authfilenumberna("not_found"),role(role_type),source(row),altrender(altrender_type),tid(row,arg))
                full_text.append(text)
        return full_text
    else:
        return ""
def type1(arg,header_row):
    label_title=header_row[arg]
    return{"type":label_title}

#The actual code starts here: the input should be the name of the spss to gather.
#One tab each shelfmark to gather. 

wb_input = input ('Please write the name of the spreadsheet to Gather: ')
wb_name = wb_input + '.xlsx'
wb = load_workbook(wb_name, read_only=True)
shelfmarks = wb.sheetnames
for shelfmark_modified in shelfmarks:
    rec_num = 1
    tid_num=1
    try:
        ws = wb[shelfmark_modified]
    except KeyError:
        print("Sheet not found")


#This part defines where the authority files details are held.

    auth_file_name = 'Authorities_combined2.xlsx'
    auth_file_wb = load_workbook(auth_file_name, read_only=True)
    try:
        auth_ws = auth_file_wb["1"]
    except KeyError:
        print("Sheet not found")

#This is where the nodes and dependencies are established. If a node has teh wrng name, change it in E.node
        
    E = ElementMaker(namespace="urn:isbn:1-931666-22-9",nsmap={'ead': "urn:isbn:1-931666-22-9", 'xlink': "http://www.w3.org/1999/xlink", 'xsi': "http://www.w3.org/2001/XMLSchema-instance"})
    
    EAD = E.ead
    EADHEADER = E.eadheader
    EADID = E.eadid
    FILEDESC = E.filedesc
    TITLESTMT = E.titlestmt
    TITLEPROPER = E.titleproper
    PROFILEDESC = E.profiledesc
    CREATION = E.creation
    DATE = E.date
    LANGUSAGE = E.langusage
    LANGUAGE = E.language
    #LANGUAGE1 = E.language1
    ARCHDESC = E.archdesc
    DID = E.did
    REPOSITORY = E.repository
    UNITID = E.unitid
    UNITTITLE = E.unittitle
    TITLE = E.title
    UNITDATE = E.unitdate
    LANGMATERIAL = E.langmaterial
    PHYSDESC = E.physdesc
    EXTENT = E.extent
    ACCESSRESTRICT = E.accessrestrict
    P = E.p
    LEGALSTATUS = E.legalstatus
    ACCRUALS = E.accruals
    BIOGHIST = E.bioghist
    APPRAISAL = E.appraisal
    ARRANGEMENT = E.arrangement
    PHYSTECH = E.phystech
    SCOPECONTENT = E.scopecontent
    LIST = E.list
    USERRESTRICT = E.userrestrict
    ODD = E.odd
    CONTROLACCESS = E.controlaccess
    GENREFORM = E.genreform
    PERSNAME = E.persname
    FAMNAME = E.famname
    CORPNAME = E.corpname
    SUBJECT = E.subject
    GEOGNAME = E.geogname
    NOTE = E.note
    
    full_ead = EAD()
    header_row = get_header()


# This bit creates the parts of tree for each child shelfmark.

    for row in ws.iter_rows(min_row=2, values_only=False):
       ead = EAD()
       comment = Comment(f"New record starts here {row[5].value}")
       full_ead.append(comment)
       shelfmark = str(row[5].value)       
       print(shelfmark)
       
       eadheader = EADHEADER(StartRecord(str(rec_num)))
       eadid = EADID(str(shelfmark),tid(row,5))
       filedesc = FILEDESC()
       titlestmt = TITLESTMT()
       titleproper = TITLEPROPER()
       profiledesc = PROFILEDESC()
       creation = CREATION()
       date = DATE()
       date = DATE(str(datetime.now().strftime("%Y-%m-%dT%H:%M:%S")), date_type("exported"),tid(row,5))
       creation.append(date)
       date = DATE(str(wb.properties.modified.strftime("%Y-%m-%dT%H:%M:%S")), date_type("modified"),tid(row,5))
       creation.append(date)
       langusage = LANGUSAGE()
       language = LANGUAGE(content(row,40),langcode(row),scriptcode(row),tid(row,40))
       
       
       archdesc = ARCHDESC(level(row))
       did = DID()
       repository= REPOSITORY(row[0].value + ": " + row[1].value,tid(row,0))
       unitid = UNITID(shelfmark, IAMSlabel(row), identifier(row),tid(row,5))
       unittitle1 = UNITTITLE(label(10,header_row))
       title = TITLE(content(row,10),tid(row,10))
       unittitle2 = UNITTITLE(content(row,7),label(7,header_row),tid(row,7))
       unittitle3 = UNITTITLE(content(row,6),label(6,header_row),tid(row,6))
       unitdate = UNITDATE(date_in_full(row),datechar(row),calendar(row),era(row),normal(row),tid(row,14))
       langmaterial = LANGMATERIAL()
       languages = row[40].value.split("|")
       lang_codes = row[41].value.split("|")
       for i in range(0,len(languages)):
           l = languages[i]
           c = lang_codes[i]
           language1 = LANGUAGE(lang_content(l),mat_langcode(c),tid(row,41))
           langmaterial.append(language1)
                  
       langmaterial1 = LANGMATERIAL()
       language2 = LANGUAGE(content(row,42),mat_scriptcode(row),tid(row,43))
       physdesc = PHYSDESC()
       extent = EXTENT(content(row,19),tid(row,19))
       
       
       accessrestrict = ACCESSRESTRICT()
       for p in pcontent(row, 25):
           accessrestrict.append(p)
       accessrestrict1 = ACCESSRESTRICT()
       legalstatus = LEGALSTATUS(content(row,71),tid(row,71))
       accruals = ACCRUALS()
       for p in pcontent(row, 23):
           accruals.append(p)
       bioghist = BIOGHIST()
       for p in pcontent(row, 24):
           bioghist.append(p)
       appraisal = APPRAISAL()
       for p in pcontent(row, 22):
           appraisal.append(p)
       arrangement = ARRANGEMENT()
       for p in pcontent(row,31):
           arrangement.append(p)
       phystech = PHYSTECH()
       for p in pcontent(row,21):
           phystech.append(p)
      
       scopecontent = SCOPECONTENT()
       lists = LIST()
       if row[20].value.find("-"):
           list_content = []
           top_content = []
           bottom_content = []
           lines = row[20].value.split("\n")
           for line in lines:
               if line.startswith("-"):
                  list_content.append(line)
               else:
                  if list_content == []:
                      top_content.append(line)
                  else:
                      bottom_content.append(line)
           for section in top_content:
               p = E.p(section, tid(row,20))   
               scopecontent.append(p)
           for section in list_content:
               item = E.item(section.strip("-"), tid(row, 20))  
               lists.append(item)
               scopecontent.append(lists)
           for section in bottom_content:
               p = E.p(section, tid(row,20))   
               scopecontent.append(p)
       else:
            for p in pcontent(row,20):
                scopecontent.append(p)               
       userestrict = USERRESTRICT()
       for p in pcontent(row, 27):
           userestrict.append(p)
       odd = ODD()
       for p in pcontent(row, 36):
           #unsure if this is the correct mapping for this field!
           odd.append(p)
       controlaccess = CONTROLACCESS()
       genreform = GENREFORM(content(row,79),source(row),tid(row,79))
       controlaccess.append(genreform)
       
       
#Authority files processing starts here:
    
       print(str(datetime.now().strftime("%Y-%m-%dT%H:%M:%S"))+' processing authority files for ' + shelfmark +'...')
       for arg in range(48,54,1):
           for authorityfile in authorityfiles(row,arg):
               controlaccess.append(authorityfile)
       note1 = NOTE(type1(2,header_row))
       for p in pcontent(row,1):
           note1.append(p)
       note2 = NOTE(type1(2,header_row))
       for p in pcontent(row,2):
           note2.append(p)
           
           
#Tree structure starts here:
    
    #Header creation 
       ead.append(eadheader)
       eadheader.append(eadid)
       eadheader.append(filedesc)
       filedesc.append(titlestmt)
       titlestmt.append(titleproper)
       eadheader.append(profiledesc)
       profiledesc.append(creation)
       creation.append(date)
       creation.append(date)
       profiledesc.append(langusage)
       langusage.append(language)
    #Archdesc creation
       ead.append(archdesc)
       archdesc.append(did)
       did.append(repository)
       did.append(unitid)
       did.append(unittitle1)
       unittitle1.append(title)
       did.append(unittitle2)
       did.append(unittitle3)
       did.append(unitdate)
       did.append(langmaterial)
       
       did.append(langmaterial1)
       langmaterial1.append(language2)
       did.append(physdesc)
       physdesc.append(extent)
       archdesc.append(accessrestrict)
       archdesc.append(accessrestrict1)
       accessrestrict1.append(legalstatus)
       archdesc.append(accruals)
       archdesc.append(bioghist)
       archdesc.append(appraisal)
       archdesc.append(arrangement)
       archdesc.append(phystech)
       archdesc.append(scopecontent)
       archdesc.append(userestrict)
       archdesc.append(odd)
       archdesc.append(controlaccess)
       controlaccess.append(note1)
       controlaccess.append(note2)
       
       rec_num = rec_num+1
 
#This part puts together the two parts of each child's (and parent's) tree (header+description)
#This will append as many children as there are in the Excel tab       
       full_ead.append(eadheader)
       full_ead.append(archdesc)   
       
#This part writes out the XML file.
    with open(shelfmark_modified+'.xml', 'wb') as f:
        f.write(etree.tostring(full_ead, encoding='UTF-8', pretty_print=True))
        
    print(shelfmark + ' complete \n')
print('Gather complete!')